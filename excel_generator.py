"""Модуль для генерации Excel-файла с извещением на оплату."""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from models import PaymentNotice
from date_time_utils import DateTimeHandler


class ExcelGenerator:
    """Класс для генерации Excel-файла с извещением на оплату."""

    def __init__(self):
        """Инициализирует генератор Excel."""
        self._date_handler = DateTimeHandler()

    def generate_payment_notice(self, notice: PaymentNotice, file_path: str) -> None:
        """
        Генерирует Excel-файл с извещением на оплату.

        Args:
            notice: Объект извещения на оплату.
            file_path: Путь для сохранения файла.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Извещение на оплату"

        # Стили
        header_font = Font(bold=True, size=14, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        title_font = Font(bold=True, size=16)
        label_font = Font(bold=True)
        border_side = Side(style='thin', color="000000")
        border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

        # Заголовок
        ws.merge_cells('A1:D1')
        header_cell = ws['A1']
        header_cell.value = "ИЗВЕЩЕНИЕ НА ОПЛАТУ"
        header_cell.font = header_font
        header_cell.fill = header_fill
        header_cell.alignment = Alignment(horizontal='center', vertical='center')

        # Информация о лицевом счете
        row = 3
        ws[f'A{row}'] = "Лицевой счет:"
        ws[f'B{row}'] = notice.account.account_number
        ws[f'A{row}'].font = label_font

        row += 1
        ws[f'A{row}'] = "ФИО:"
        ws[f'B{row}'] = notice.account.full_name
        ws[f'A{row}'].font = label_font

        row += 1
        ws[f'A{row}'] = "Адрес:"
        ws[f'B{row}'] = notice.account.get_address(notice.street)
        ws[f'A{row}'].font = label_font

        row += 1
        period_name = self._date_handler.get_month_name(notice.period_month)
        ws[f'A{row}'] = "Период:"
        ws[f'B{row}'] = f"{period_name} {notice.period_year}"
        ws[f'A{row}'].font = label_font

        row += 1
        ws[f'A{row}'] = "Дата формирования:"
        ws[f'B{row}'] = self._date_handler.format_date(self._date_handler.get_current_date())
        ws[f'A{row}'].font = label_font

        # Таблица начислений
        row += 2
        headers = ["№", "Услуга", "Количество", "Тариф", "Сумма"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # Данные начислений
        row += 1
        for idx, (charge, service) in enumerate(notice.charges, start=1):
            ws.cell(row=row, column=1, value=idx).border = border
            ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')

            ws.cell(row=row, column=2, value=service.name).border = border
            ws.cell(row=row, column=3, value=charge.quantity).border = border
            ws.cell(row=row, column=3).alignment = Alignment(horizontal='right')
            ws.cell(row=row, column=4, value=service.tariff).border = border
            ws.cell(row=row, column=4).alignment = Alignment(horizontal='right')
            
            cost = service.calculate_cost(charge.quantity)
            ws.cell(row=row, column=5, value=cost).border = border
            ws.cell(row=row, column=5).alignment = Alignment(horizontal='right')
            row += 1

        # Итого
        ws.merge_cells(f'A{row}:D{row}')
        total_label_cell = ws[f'A{row}']
        total_label_cell.value = "ИТОГО К ОПЛАТЕ:"
        total_label_cell.font = Font(bold=True, size=12)
        total_label_cell.alignment = Alignment(horizontal='right', vertical='center')
        total_label_cell.border = border

        total_cell = ws[f'E{row}']
        total_cell.value = notice.total_amount
        total_cell.font = Font(bold=True, size=12)
        total_cell.alignment = Alignment(horizontal='right', vertical='center')
        total_cell.border = border
        total_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        # Настройка ширины столбцов
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15

        # Сохранение файла
        wb.save(file_path)

